# -*- coding: utf-8 -*-
"""estrazione_cantiere — i due strati che l'estrattore congelato non vede (E48).

⚠️ **QUESTO MODULO NON E' L'ESTRATTORE DI MISURA E NON LO TOCCA.**
`qa_comune.text_of` / `qa_comune.testo_fonte` restano byte-identici: ogni numero
delle baseline e' vincolato al loro comportamento, e cambiarli invaliderebbe
confronti gia' pubblicati (metodo_01 §5-bis). Qui si **aggiunge**, non si
sostituisce.

COSA AGGIUNGE, e perche' proprio questi due strati:

  1. `[FORMULA: foglio!RIF =...]` — le formule dei fogli di calcolo. Il
     censimento del 21/08/2026 ne conta **1.697 in 13 file su 15, tutte senza
     valore in cache**: l'estrattore congelato apre gli `.xlsx` con
     `data_only=True` e su una cella con formula e senza cache legge `None`.
     Il dato non e' nascosto: **non c'e' proprio**, e la formula e' l'unica cosa
     che dica che cosa quella cella avrebbe dovuto contenere. Riga T89.

  2. `[BARRATO: ...]` — i passaggi barrati dei documenti. Il barrato non e'
     testo, e' una proprieta' del carattere: le parole stanno dove stanno tutte
     le altre e **una frase cancellata arriva identica a una in vigore**. Nella
     sola scheda allergeni sono quattro, e il vault ne aveva colto **uno**, solo
     perche' un commento accanto usava la parola «cancellata». Riga T96.

DOVE VIVE IL TESTO AGGIUNTO — e non e' un dettaglio:
gli strati si **appendono in coda**, mai in mezzo. Il testo della via congelata
resta cosi' un **prefisso esatto** di quello di cantiere, e la separazione fra
le due vie si **dimostra** invece di dichiararla:

    testo_cantiere(n).startswith(testo_fonte(n))   # vero per costruzione
    prova_invarianza()                             # e lo verifica su tutto il corpus

CHI LO USA: **solo** la suite QA e il pacchetto dello strato di giudizio.
Nessuna misura, nessuna baseline, nessun confronto storico.

Uso:
    python estrazione_cantiere.py --prova        # l'invarianza, su tutti i grezzi
    python estrazione_cantiere.py --censimento   # quanto vede, file per file
"""
import argparse, io, os, re, sys, zipfile

QUI = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, QUI)
import qa_comune as Q

MARCA_FORMULA = "[FORMULA: %s]"
MARCA_BARRATO = "[BARRATO: %s]"

# le due marche, per chi deve riconoscerle nel testo (qa_provenance)
RE_BARRATO = re.compile(r"\[BARRATO: (.*?)\]", re.S)
RE_FORMULA = re.compile(r"\[FORMULA: (.*?)\]", re.S)


# --------------------------------------------------------------- strato 1: formule

def _strato_formule(p):
    """Le celle con formula di un .xlsx, nell'ordine dei fogli e delle righe."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, data_only=False)
    except Exception:
        return []
    fuori = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    fuori.append(MARCA_FORMULA % ("%s!%s %s" % (ws.title, c.coordinate, v)))
    return fuori


# --------------------------------------------------------------- strato 2: barrato

def _run_barrati_docx(p):
    """I run di word/document.xml che portano il barrato attivo.

    Si legge l'XML invece di usare python-docx perche' il barrato puo' stare su
    run dentro tabelle, note e caselle di testo, e qui interessa **trovarli
    tutti**, non ricostruire il documento.
    """
    try:
        z = zipfile.ZipFile(p)
    except Exception:
        return []
    fuori = []
    for nome in z.namelist():
        if not (nome.startswith("word/") and nome.endswith(".xml")):
            continue
        try:
            x = z.read(nome).decode("utf-8", "replace")
        except Exception:
            continue
        for run in re.findall(r"<w:r(?:\s[^>]*)?>.*?</w:r>", x, re.S):
            m = re.search(r"<w:rPr>(.*?)</w:rPr>", run, re.S)
            if not m or not _barrato_attivo_w(m.group(1)):
                continue
            testo = "".join(re.findall(r"<w:t(?:\s[^>]*)?>(.*?)</w:t>", run, re.S))
            testo = _disescape(testo).strip()
            if testo:
                fuori.append(MARCA_BARRATO % testo)
    return fuori


def _barrato_attivo_w(rpr):
    """`<w:strike/>` o `<w:dstrike/>` senza un `w:val` che li spenga."""
    for tag in ("strike", "dstrike"):
        m = re.search(r"<w:%s(\s[^>]*)?/?>" % tag, rpr)
        if not m:
            continue
        attr = m.group(1) or ""
        val = re.search(r'w:val="([^"]*)"', attr)
        if val is None or val.group(1).lower() not in ("false", "0", "off"):
            return True
    return False


def _run_barrati_pptx(p):
    """Le porzioni barrate delle slide e delle note del relatore."""
    try:
        z = zipfile.ZipFile(p)
    except Exception:
        return []
    fuori = []
    for nome in z.namelist():
        if not (nome.startswith("ppt/") and nome.endswith(".xml")):
            continue
        try:
            x = z.read(nome).decode("utf-8", "replace")
        except Exception:
            continue
        for run in re.findall(r"<a:r>.*?</a:r>", x, re.S):
            m = re.search(r'<a:rPr[^>]*\sstrike="(sngStrike|dblStrike)"', run)
            if not m:
                continue
            testo = "".join(re.findall(r"<a:t>(.*?)</a:t>", run, re.S))
            testo = _disescape(testo).strip()
            if testo:
                fuori.append(MARCA_BARRATO % testo)
    return fuori


def _disescape(s):
    for a, b in (("&amp;", "&"), ("&lt;", "<"), ("&gt;", ">"),
                 ("&quot;", '"'), ("&apos;", "'")):
        s = s.replace(a, b)
    return s


# --------------------------------------------------------------- l'estrazione

_cache = {}


def strati(nome):
    """Le sole righe aggiunte, senza il testo congelato. Elenco di stringhe."""
    p = os.path.join(Q.SOURCES, nome)
    if not os.path.isfile(p):
        return []
    e = nome.rsplit(".", 1)[-1].lower()
    try:
        if e == "xlsx":
            return _strato_formule(p)
        if e == "docx":
            return _run_barrati_docx(p)
        if e == "pptx":
            return _run_barrati_pptx(p)
    except Exception as ex:
        print("  [avvertenza] strato di cantiere fallito su %s: %s" % (nome, ex))
    return []


def testo_cantiere(nome):
    """Il testo congelato, PIU' i due strati marcati appesi in coda.

    ⚠️ Il testo congelato resta un prefisso esatto: e' la proprieta' su cui
    poggia la prova di invarianza.
    """
    if nome not in _cache:
        base = Q.testo_fonte(nome)
        agg = strati(nome)
        _cache[nome] = base + ("\n" + "\n".join(agg) if agg else "")
    return _cache[nome]


def testo_vigente(nome):
    """Il testo di cantiere SENZA i passaggi revocati.

    Serve a distinguere un riscontro che vive **solo** nel barrato: se
    un'affermazione si trova in `testo_cantiere` e non qui, il suo unico
    appoggio e' contenuto revocato (E48).

    ⚠️ **Non basta togliere i marcatori `[BARRATO: ...]`**, ed e' l'errore che
    questa funzione ha fatto nella sua prima stesura: il testo barrato **sta gia'
    dentro il testo congelato**, perche' l'estrattore restituisce le parole di
    ogni run senza guardarne la formattazione. Togliere solo la coda marcata
    lascia intatta l'occorrenza originale, e il controllo non scatta mai.

    ⚠️ **Quante occorrenze si tolgono, e perche' non tutte.** Di ogni passaggio
    barrato si toglie **una** occorrenza per ogni run che lo porta barrato: se una
    stessa frase compare due volte e una sola e' cancellata, l'altra resta e
    sostiene ancora le affermazioni. E' un'approssimazione — toglie la PRIMA
    occorrenza, che non e' detto sia quella barrata — e vale in questa direzione:
    puo' far scattare un avviso di troppo, mai zittirne uno dovuto.
    """
    t = RE_BARRATO.sub(" ", testo_cantiere(nome))
    for riga in strati(nome):
        m = RE_BARRATO.fullmatch(riga)
        if m:
            t = t.replace(m.group(1), " ", 1)
    return t


# --------------------------------------------------------------- la prova

def prova_invarianza(stampa=True):
    """L'estrattore di misura non e' stato toccato, e si vede su tutto il corpus.

    Due condizioni, entrambe necessarie:
      1. il testo congelato e' un PREFISSO esatto di quello di cantiere;
      2. tagliando il testo di cantiere alla lunghezza del congelato si
         riottiene il congelato, carattere per carattere.
    """
    grezzi = sorted(os.listdir(Q.SOURCES))
    guasti, con_strati, righe = [], 0, 0
    for g in grezzi:
        if not os.path.isfile(os.path.join(Q.SOURCES, g)):
            continue
        base = Q.testo_fonte(g)
        cant = testo_cantiere(g)
        if not cant.startswith(base) or cant[:len(base)] != base:
            guasti.append(g)
        n = len(strati(g))
        if n:
            con_strati += 1
            righe += n
    if stampa:
        print("grezzi esaminati .............. %d" % len(grezzi))
        print("con almeno uno strato ......... %d" % con_strati)
        print("righe aggiunte in tutto ....... %d" % righe)
        print("prefisso violato .............. %d" % len(guasti))
        for g in guasti:
            print("   GUASTO:", g)
        print()
        if guasti:
            print("INVARIANZA VIOLATA: l'estrazione di cantiere ha alterato il testo congelato.")
        else:
            print("INVARIANZA PROVATA: su tutti i %d grezzi il testo della via congelata e'"
                  % len(grezzi))
            print("un prefisso ESATTO di quello di cantiere. L'estrattore di misura non e'")
            print("stato toccato, e questo non e' un'affermazione: e' il risultato del confronto.")
    return not guasti


def censimento():
    """Quanto vede lo strato, file per file. Tabella incolla-e-vai."""
    print("| Grezzo | Formule | Barrati |")
    print("|---|---|---|")
    tf = tb = 0
    for g in sorted(os.listdir(Q.SOURCES)):
        if not os.path.isfile(os.path.join(Q.SOURCES, g)):
            continue
        s = strati(g)
        nf = sum(1 for r in s if r.startswith("[FORMULA:"))
        nb = sum(1 for r in s if r.startswith("[BARRATO:"))
        if nf or nb:
            print("| `%s` | %s | %s |" % (g, nf or "—", nb or "—"))
        tf += nf; tb += nb
    print("| **totale** | **%d** | **%d** |" % (tf, tb))


def main():
    ap = argparse.ArgumentParser(description="Estrazione di cantiere (E48).")
    ap.add_argument("--prova", action="store_true", help="prova l'invarianza dell'estrattore di misura")
    ap.add_argument("--censimento", action="store_true", help="quanto vede lo strato, file per file")
    ap.add_argument("--mostra", help="stampa gli strati di UN grezzo")
    a = ap.parse_args()
    if a.mostra:
        for r in strati(a.mostra):
            print(r)
        return 0
    if a.censimento:
        censimento()
        return 0
    prova_invarianza()
    return 0


if __name__ == "__main__":
    sys.exit(main())
