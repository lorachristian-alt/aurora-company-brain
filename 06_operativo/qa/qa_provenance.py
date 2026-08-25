# -*- coding: utf-8 -*-
"""qa_provenance — ogni fatto affermato ha riscontro nella fonte citata.

metodo_03 §7.1. E' il controllo che vale piu' di tutti gli altri messi insieme,
ed e' il motivo per cui esiste la sezione `## Fonti` con i locator.

DUE STRATI, che non si confondono mai (un controllo che li mescola non e'
riproducibile, e un numero non riproducibile non vale — regola d'oro 5):

  * strato DETERMINISTICO, qui dentro, senza LLM: estrae dal corpo le stringhe
    verificabili con una regex — numeri, date, orari, codici, citazioni fra
    virgolette basse — e le cerca nel testo delle fonti, dopo normalizzazione
    alias. E' questo strato a produrre gli ERRORI.
    ⚠️ **Dal 21/08/2026 (E48) il riscontro si cerca nell'ESTRAZIONE DI CANTIERE**
    (`estrazione_cantiere.testo_cantiere`), che parte da `text_of` congelata e vi
    APPENDE marcati i due strati che quella non vede: le formule dei fogli di
    calcolo e i passaggi barrati dei documenti. **L'estrattore di misura resta
    byte-identico** e la separazione si prova (`estrazione_cantiere --prova`).

  * strato di GIUDIZIO, che gira altrove (subagente a contesto pulito): risponde
    alle due domande che una regex non puo' porre. Produce solo AVVISI. Il testo
    del prompt sta qui sotto in PROMPT_GIUDIZIO, congelato con la sua data:
    da li' in poi ogni lotto viene giudicato con le stesse parole.

Uso:
    python qa_provenance.py --perimetro lotto @fetta_l26130.txt
    python qa_provenance.py --perimetro lotto @fetta.txt --pacchetto-giudizio
"""
import argparse, io, os, re, sys
from datetime import date, datetime

import qa_comune as Q
import estrazione_cantiere as EC

CONTROLLO = "provenance"

# --------------------------------------------------------------------------
# PROMPT DI GIUDIZIO — congelato il 2026-08-16 (Sessione 2, gate del pilota).
# Deriva dalla disciplina del giudice P3 di metodo_02: rubrica chiusa, una riga
# JSON per nota, nessuna riscrittura, niente generosita'.
# Non si modifica per un lotto: si modifica per tutti, con una data nuova.
# --------------------------------------------------------------------------
# La rete interna di QA evolve con versioni DICHIARATE, come metodo_03. Il congelamento
# intoccabile riguarda gli strumenti di MISURA (P1, P3, config C), dove la confrontabilita'
# prima/dopo e' il prodotto. Ogni rapporto di lotto dichiara la versione usata: il lotto 1A
# e' stato giudicato con la v1, dal lotto 1B vale la v2. Mai retroattiva.
PROMPT_GIUDIZIO_VERSIONE = "v2"
PROMPT_GIUDIZIO_DATA = "2026-08-18"          # v1: 2026-08-16
PROMPT_GIUDIZIO = """\
Giudichi note di un archivio aziendale rispetto ai documenti che dichiarano come
fonte. Non devi migliorare le note, non devi riscriverle e non devi rispondere
alle domande che pongono: devi giudicare se stanno in piedi sulle loro fonti.

Per ogni nota ricevi: il nome, il corpo integrale, l'elenco delle fonti che
dichiara e il testo estratto di ciascuna di quelle fonti.

Rispondi SOLO a queste tre domande, in quest'ordine:
1. Ogni fonte elencata contribuisce davvero alla nota? Una fonte che non sorregge
   nessuna affermazione e' rumore nel payload, e va segnalata.
2. La nota afferma qualcosa che le fonti non dicono, pur senza numeri? Cerca le
   affermazioni qualitative: cause, intenzioni, conseguenze, attribuzioni di una
   frase a una persona, giudizi di conformita'. I numeri, le date e i codici NON
   sono affar tuo: li ha gia' verificati lo strato deterministico.

3. Esiste, FRA LE FONTI CHE HAI RICEVUTO IN QUESTO PACCHETTO, un documento che
   misura o afferma la stessa grandezza di cui la nota parla, e che la nota NON
   cita? Non e' un difetto di provenienza: e' una lacuna di copertura. La segnali
   FUORI dal verdetto, in coda alla risposta, dicendo quale nota e quale
   documento, e NON cambi per questo l'esito della nota.

Assegna a ogni nota esattamente un esito:
- `pulita` — ogni fonte contribuisce e nessuna affermazione eccede le fonti
- `fonte_inutile` — almeno una fonte elencata non sorregge nulla
- `afferma_oltre` — almeno un'affermazione non ha riscontro nelle fonti citate
- `entrambi` — sia una fonte inutile sia un'affermazione che eccede

Restituisci UNA riga JSON per nota, con esattamente questi campi:
`nota`, `esito`, `motivazione` (una riga sola, che cita l'affermazione o la fonte
esatta a cui ti riferisci).

Regole che non puoi violare:
- Non essere generoso: se un'affermazione non e' ritrovabile leggendo le fonti,
  non e' fondata anche se e' plausibile e ben scritta.
- Un'inferenza DICHIARATA come tale («i due documenti, letti insieme, mostrano
  che…», «la trascrizione non lo afferma: e' un'attribuzione») e' legittima e non
  si segnala. E' l'inferenza fatta passare per dato della fonte che si segnala.
- Una nota che riporta due valori divergenti senza sceglierne uno sta facendo il
  suo mestiere: non e' un difetto, e non si segnala.
- Non proponi correzioni e non riscrivi niente. Riporti, e basta.
- Non ricevi il canone e non devi cercarlo: confronti la nota contro le sue
  fonti, punto. Il giudizio sulla copertura dei fatti spetta a un altro ruolo.
- La terza domanda non ti autorizza a giudicare la copertura dell'archivio: ti
  chiede solo di segnalare cio' che hai gia' sotto gli occhi dentro il pacchetto.
  Se un documento non e' nel pacchetto, per te non esiste.
"""

# ------------------------------------------------------------------ estrazione

MESI = "gen|feb|mar|apr|mag|giu|lug|ago|set|ott|nov|dic"

RE_CITAZIONE = re.compile(r"«([^»]{4,})»")
RE_DATA = re.compile(r"\b(?:\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}|\d{1,2}-(?:%s)-\d{2,4})\b" % MESI, re.I)
RE_ORA = re.compile(r"\b\d{1,2}:\d{2}(?::\d{2})?\b")
RE_CODICE = re.compile(
    r"\b(?:AF-[A-Z]{2}-\d{4}|L26\d{3}(?:-L\d-T\d)?|MV26[-_]\d{4}[/-]?[A-Z]|"
    r"MOD-[A-Z]{2}-\d{2}|NC-\d{4}-\d{3}|REC-\d{4}-\d{3}|IO-\d{2}|CCP\d|"
    r"PT-?\d{3}|PKM-?\d{3}|MD-?\d{4}|CIP-?\d{2}|AL-\d{3}|E-\d{3}|"
    r"BRC/[A-Z]{2}/\d{2}/\d{5}|\d{13,14})\b")
# numeri "che contano": con separatore, con decimale, oppure da tre cifre in su.
RE_NUMERO = re.compile(r"(?<![\w.,/:-])\d{1,3}(?:\.\d{3})+(?:,\d+)?(?![\w/])"
                       r"|(?<![\w.,/:-])\d+,\d+(?![\w/])"
                       r"|(?<![\w.,/:-])\d{3,}(?![\w.,/:-])")

# marcatori che dichiarano un valore come derivato (metodo_03 §5.4)
RE_DERIVATO = re.compile(r"\((?:calcolat|contat|derivat|somma|differenza)[^)]*\)", re.I)
RE_FORMULA = re.compile(r"(\d[\d.,]*(?:\s*[+\-]\s*\d[\d.,]*)+)\s*=\s*\*{0,2}(\d[\d.,]*)")


# Una «...» nel corpo non e' sempre una citazione: in italiano le virgolette basse
# marcano anche il nome di un foglio di calcolo, il titolo di una sezione, un'etichetta di
# colonna — e metodo_03 le usa cosi' nei propri esempi (`foglio «A valle»`). Pretendere il
# riscontro testuale su tutte boccia le note corrette. Soglia: sotto le cinque parole non
# e' una citazione, e' un nome.
PAROLE_MIN_CITAZIONE = 5


def e_citazione(testo):
    # si contano solo i token con almeno un carattere alfanumerico: un trattino isolato
    # dentro il titolo di una sezione non e' una parola, e farlo contare fa scattare la
    # soglia su nomi di sezione che citazioni non sono.
    parole = [p for p in testo.split() if any(c.isalnum() for c in p)]
    return len(parole) >= PAROLE_MIN_CITAZIONE


def ripulisci_citazione(s):
    """Toglie l'enfasi markdown: nel grezzo non c'e', e non deve far fallire il confronto.

    ⚠️ Si tolgono asterischi e backtick, NON gli underscore: `T_CUORE`, `TT_02`, `MV26_0429A`
    sono codici veri di questo corpus, e mangiarne l'underscore fa fallire il confronto
    proprio sulle citazioni piu' precise."""
    return re.sub(r"[*`]+", "", s)


def varianti_data(s):
    """Il corpus scrive le date in tre formati (metodo_01 §2): la stessa data va cercata
    con l'anno a quattro e a due cifre, altrimenti una nota che scrive 10/05/2026 fallisce
    contro un foglio che scrive 10/05/26."""
    v = {s}
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", s)
    if m:
        g, me, a = m.groups()
        v.add("%s/%s/%s" % (g, me, a[-2:]))
        v.add("%s/%s/20%s" % (g, me, a[-2:]))
        v.add("%02d/%02d/%s" % (int(g), int(me), a[-2:]))
        v.add("%02d/%02d/20%s" % (int(g), int(me), a[-2:]))
        v.add("20%s-%02d-%02d" % (a[-2:], int(me), int(g)))
    return v


def varianti_numero(s):
    """Le scritture sotto cui lo stesso numero puo' comparire in un grezzo:
    separatori invertiti, senza separatori, senza decimali nulli."""
    v = {s}
    v.add(s.replace(".", "").replace(",", "."))
    v.add(s.replace(".", "X").replace(",", ".").replace("X", ","))
    v.add(s.replace(".", ""))
    v.add(s.replace(".", "").replace(",", ""))
    if re.match(r"^\d+,0+$", s):
        v.add(s.split(",")[0])
    if "," in s:
        v.add(s.replace(",", "."))
    return {x for x in v if x}


# ⚠️ LA SUPERFICIE `title`/`summary` ENTRA NEL CONTROLLO IL 23/08/2026, E UN CONTROLLO
# NUOVO NON RENDE ROSSO IL PREGRESSO (§4.35). Al momento in cui entra, il vault porta
# quattordici affermazioni verificabili che vivono solo nell'intestazione e che nessuna fonte
# citata sorregge: quasi tutte sono date scritte con l'anno dove la fonte non lo scrive
# (E24 - la grafia della fonte; E50 - l'anno inferito e' un derivato). Pretenderlo da note
# nate quando nessuno guardava li' bloccherebbe ogni lotto futuro dietro una sanatoria.
# Quindi: ERRORE per le note nate con la regola in vigore, AVVISO DICHIARATO per il pregresso
# - la stessa disciplina con cui E43 e' entrato il 20/08 (`qa_frontmatter.NASCITA_E43`).
NASCITA_SUPERFICIE_INTESTAZIONE = date(2026, 8, 23)


def nota_e_nuova(nota, nascita):
    """Vero se la nota e' nata con la regola gia' in vigore. `data_nota` e' il padrone."""
    dn = (nota.fm or {}).get("data_nota")
    try:
        return datetime.strptime(str(dn), "%Y-%m-%d").date() >= nascita
    except (ValueError, TypeError):
        return isinstance(dn, date) and dn >= nascita


def estrai_affermazioni(testo):
    """Le stringhe che una regex puo' verificare. Ritorna [(genere, testo)]."""
    fuori = []
    # si tolgono i wikilink: un nome di nota non e' un'affermazione sul mondo
    pulito = re.sub(r"\[\[[^\]]+\]\]", " ", testo)
    for m in RE_CITAZIONE.finditer(pulito):
        t = m.group(1).strip()
        if e_citazione(t):
            fuori.append(("citazione", ripulisci_citazione(t)))
    senza_cit = RE_CITAZIONE.sub(" ", pulito)
    for genere, rx in (("codice", RE_CODICE), ("data", RE_DATA),
                       ("ora", RE_ORA), ("numero", RE_NUMERO)):
        for m in rx.finditer(senza_cit):
            fuori.append((genere, m.group(0)))
        senza_cit = rx.sub(" ", senza_cit)
    return fuori


def numeri_esenti(testo):
    """I valori derivati DICHIARATI: il risultato di una formula scritta per
    esteso, o un numero marcato come calcolato/contato. La QA verifica gli
    addendi, non il risultato (§5.4 e §7.1 clausola 2)."""
    esenti = set()
    for m in RE_FORMULA.finditer(testo):
        esenti |= varianti_numero(m.group(2))
    for m in RE_DERIVATO.finditer(testo):
        inizio = max(0, m.start() - 60)
        for n in RE_NUMERO.finditer(testo[inizio:m.start()]):
            esenti |= varianti_numero(n.group(0))
    return esenti


# -------------------------------------------------------------- verifica nota

# ⚠️ E48 — QUANDO UN RISCONTRO VIVE SOLO IN TESTO BARRATO.
# Il barrato e' contenuto REVOCATO: chi lo cita come se fosse in vigore afferma il
# falso, ed e' la classe piu' grave del progetto. Ma una nota che PARLA del barrato
# — e ce ne sono, ed e' esattamente quello che deve fare — riporta per forza le
# parole revocate, e segnalarla sarebbe punire il comportamento giusto.
#
# Il criterio e' a livello di NOTA e non di singola citazione, ed e' una scelta:
# una nota che dice da qualche parte «barrato», «revocato» o «cancellato» sa di
# star maneggiando testo cancellato, e la distinzione fine fra le sue frasi non e'
# alla portata di una regex. ⚠️ **Il collaudo pianta entrambi i versi**: una nota
# che afferma come vigente un testo barrato (deve essere segnalata) e una che lo
# dichiara revocato (non deve esserlo).
RE_DICHIARA_REVOCA = __import__("re").compile(r"barrat|revocat|cancellat|depennat", __import__("re").I)


def dichiara_la_revoca(nota):
    return bool(RE_DICHIARA_REVOCA.search(nota.grezzo))


def testo_di_riscontro(nota, per_slug):
    """Il materiale contro cui si verifica la nota.

    Clausola 4 di §7.1: hub e `_index` si verificano contro le note che elencano,
    non solo contro le proprie fonti. L'annotazione di mezza riga accanto a un
    wikilink ripete un fatto che appartiene allo spoke: il riscontro si cerca li'.
    """
    pezzi = {}
    for f in nota.fonti:
        pezzi[str(f)] = EC.testo_cantiere(str(f))
    if nota.type in ("hub", "index"):
        for target, _ in nota.wikilink():
            sp = per_slug.get(target)
            if sp is not None:
                pezzi["(spoke) " + sp.slug] = sp.corpo
    return pezzi


def controlla(nota, rep, per_slug):
    n = nota.nome
    if nota.fm is None:
        return                                   # gia' segnalato da qa_frontmatter
    if nota.type == "index" and not nota.fonti and nota.type != "hub":
        pass                                     # gli _index non hanno fonti: §2.4

    pezzi = testo_di_riscontro(nota, per_slug)
    if not pezzi:
        if nota.type not in ("index", "sessione", "daily") and not Q.e_nota_strumento(nota):
            rep.errore(n, CONTROLLO, "nessuna fonte da cui verificare le affermazioni")
        return

    norm_pezzi = {k: Q.norm(v) for k, v in pezzi.items()}
    compatti = {k: re.sub(r"[-\s/_.]", "", v) for k, v in norm_pezzi.items()}
    # lo stesso materiale, senza i passaggi revocati (E48)
    vigenti = {str(f): Q.norm(EC.testo_vigente(str(f))) for f in nota.fonti}
    comp_vig = {k: re.sub(r"[-\s/_.]", "", v) for k, v in vigenti.items()}
    nota_dichiara = dichiara_la_revoca(nota)
    ha_jpg = any(str(f).lower().endswith((".jpg", ".jpeg")) for f in nota.fonti)

    corpo = nota.corpo_senza_fonti

    # ---- LA SUPERFICIE DEL CONTROLLO: corpo PIU' intestazione -------------------------
    #
    # ⚠️ IL CONTROLLO GUARDAVA IL SOLO CORPO, E `title`/`summary` NO - riparato il
    # 23/08/2026, gate del lotto 3B, dal censimento delle superfici (§4.49). Il progetto
    # dichiara l'intestazione **portante** in due punti - E18 («se la nota stabilisce una
    # regola decisionale, il `summary` la enuncia») ed E30 («`title` e `summary` si
    # rileggono COME NOTE A SE', a ogni giro») - e non la verificava in nessuno: un numero,
    # una data o un codice inventati li' passavano la QA a **verde**.
    #
    # ⚠️ E' la superficie su cui il progetto trova piu' difetti di ogni altra: E30 nasce
    # da 1C, dove al terzo giro **sei rilievi su sette** stavano nell'intestazione col corpo
    # gia' corretto; E39, E42 ed E51 la inseguono da tre lotti; E61 nasce dal gesto che ce li
    # scrive. **Nessuno dei cinque emendamenti aveva uno strato deterministico dietro.**
    #
    # ⚠️ Il fix AGGIUNGE agganci (§4.9) e ha il suo difetto piantato in
    # `_collaudo\collaudo_intestazione.py`, o il buco si riapre in silenzio.
    #
    # Un'affermazione gia' presente nel corpo si verifica UNA volta sola: il `title` ripete
    # quasi sempre l'H1, e contarla due volte raddoppierebbe ogni rilievo senza aggiungere
    # nulla.
    intestazione = "\n".join(str((nota.fm or {}).get(k) or "") for k in ("title", "summary"))
    esenti = numeri_esenti(corpo) | numeri_esenti(intestazione)
    agganci = {k: 0 for k in pezzi}

    aff_corpo = list(estrai_affermazioni(corpo))
    gia_viste = set(aff_corpo)
    affermazioni = [(g, t, "") for g, t in aff_corpo]
    for g, t in estrai_affermazioni(intestazione):
        # una stessa affermazione compare quasi sempre due volte nell'intestazione, una
        # nel `title` e una nel `summary`: e' un rilievo solo, non due.
        if (g, t) in gia_viste:
            continue
        gia_viste.add((g, t))
        affermazioni.append((g, t, " (nell'intestazione: `title`/`summary`)"))

    for genere, tok, dove in affermazioni:
        if genere == "numero" and any(v in esenti for v in varianti_numero(tok)):
            continue
        def _cerca(nn, cc):
            if genere == "numero":
                return any(Q.presente(v, nn, cc) for v in varianti_numero(tok))
            if genere == "data":
                return any(Q.presente(v, nn, cc) for v in varianti_data(tok))
            if genere == "citazione":
                return Q.norm(tok) in nn
            return Q.presente(tok, nn, cc)

        trovato = False
        # ⚠️ nessuna uscita anticipata da questo ciclo: `agganci` serve al controllo
        # del «rumore nel payload», che conta gli agganci di OGNI fonte. Un `break`
        # qui dentro spegne le fonti successive e inventa quindici avvisi che non
        # esistono — provato il 21/08/2026 confrontando il vault prima e dopo.
        dentro_revocato = False
        dentro_vigente = False
        for k in pezzi:
            if not _cerca(norm_pezzi[k], compatti[k]):
                continue
            agganci[k] += 1
            trovato = True
            if k not in vigenti:               # gli spoke di hub e _index non hanno strati
                dentro_vigente = True
            elif _cerca(vigenti[k], comp_vig[k]):
                dentro_vigente = True
            else:
                dentro_revocato = True
        solo_revocato = dentro_revocato and not dentro_vigente
        if trovato and solo_revocato and not nota_dichiara:
            rep.avviso(n, CONTROLLO,
                       "riscontro in testo revocato: «%s» si trova SOLO in un passaggio "
                       "barrato della fonte, e la nota non lo dichiara" % tok[:60])
        if not trovato:
            msg = "%s senza riscontro in nessuna fonte citata: «%s»%s" % (genere, tok[:70], dove)
            # clausola 3: l'estrattore congelato e' cieco sulle immagini
            if ha_jpg:
                rep.avviso(n, CONTROLLO, msg + " — la nota cita un .jpg, riscontro visivo da chiudere a mano")
            elif dove and not nota_e_nuova(nota, NASCITA_SUPERFICIE_INTESTAZIONE):
                # §4.35: il pregresso dell'intestazione si dichiara debito, non si rende rosso
                rep.avviso(n, CONTROLLO,
                           msg + " — debito anteriore alla superficie dell'intestazione (23/08/2026), da sanare a fine corsa")
            else:
                rep.errore(n, CONTROLLO, msg)

    # --- le citazioni del blocco Fonti si verificano nel file di QUELLA riga -----
    for riga in nota.blocco_fonti.splitlines():
        m = re.match(r"\s*-\s*\[\[([^\]|]+)\]\]", riga)
        if not m:
            continue
        f = m.group(1).strip()
        if f.lower().endswith((".jpg", ".jpeg")):
            continue
        t = Q.norm(EC.testo_cantiere(f))
        t_vig = Q.norm(EC.testo_vigente(f))
        for cit in RE_CITAZIONE.findall(riga):
            if not e_citazione(cit):
                continue           # e' il nome di un foglio o di una sezione dentro il locator
            cit = ripulisci_citazione(cit)
            if Q.norm(cit) not in t:
                rep.errore(n, CONTROLLO,
                           "citazione non ritrovata testualmente in %s: «%s»" % (f, cit[:60]))
            else:
                agganci[f] = agganci.get(f, 0) + 1
                if Q.norm(cit) not in t_vig and not dichiara_la_revoca(nota):
                    rep.avviso(n, CONTROLLO,
                               "riscontro in testo revocato: «%s» e' barrato in %s, e la "
                               "nota non lo dichiara" % (cit[:50], f))

    # --- fix del 18/08/2026: gli identificatori marcati a codice contano come aggancio ---
    #
    # Il conteggio degli agganci si basa sulle affermazioni che una regex sa estrarre —
    # numeri, date, orari, codici di forma nota, citazioni. Una fonte puo' pero' sorreggere
    # una nota con un identificatore che la regex NON conosce: `PKM-4471-EPDM`,
    # `TST-CERT-KIT`, `PK-45.0771`. E' successo sul lotto 1A: la scheda di manutenzione
    # risultava «rumore nel payload» sulla nota dei codici del ricambio, mentre la sua riga 26
    # porta il quarto codice, che di quella nota e' il perno.
    #
    # Si contano quindi anche i token che la nota stessa marca come identificatori scrivendoli
    # fra apici inversi. ⚠️ Questo puo' solo AGGIUNGERE agganci, mai toglierne: non trasforma
    # nessun avviso in errore e non rende piu' permissivo nessun altro controllo.
    for tok in set(re.findall(r"`([A-Za-z0-9][A-Za-z0-9._/-]{3,})`", corpo)):
        for k in pezzi:
            if Q.presente(tok, norm_pezzi[k], compatti[k]):
                agganci[k] += 1

    # --- ogni fonte elencata contribuisce davvero -----------------------------------
    for f in nota.fonti:
        f = str(f)
        if f.lower().endswith((".jpg", ".jpeg")):
            rep.avviso(n, CONTROLLO,
                       "fonte immagine '%s': riscontro visivo, da chiudere a mano" % f)
            continue
        if agganci.get(f, 0) == 0:
            rep.avviso(n, CONTROLLO,
                       "la fonte '%s' non aggancia nessuna affermazione della nota: rumore nel payload" % f)

    # --- il locator punta davvero li' -------------------------------------------------
    controlla_locator_punta(nota, rep)

    # --- coerenza interna, salvo type conflitto (clausola 5) ---------------------------
    if nota.type != "conflitto":
        # ⚠️ Gli orari si tolgono PRIMA: «dalle 14:20:07 alle 14:44:37» verrebbe letto come
        # l'etichetta «dalle 14» con due valori diversi, ed e' il falso positivo piu'
        # velenoso di tutti — boccia proprio le note che descrivono bene una finestra
        # temporale, che sono quelle che contano.
        corpo_senza_orari = RE_ORA.sub(" ", corpo)
        etichette = {}
        for m in re.finditer(r"([A-Za-zà-ÿ][\w à-ÿ']{2,30}?)\s*[:=]\s*\*{0,2}(\d[\d.,]*)",
                             corpo_senza_orari):
            et, val = Q.norm(m.group(1)), m.group(2)
            etichette.setdefault(et, set()).add(val)
        for et, vals in etichette.items():
            if len(vals) > 1:
                rep.errore(n, CONTROLLO,
                           "due valori diversi per la stessa grandezza «%s»: %s"
                           % (et, ", ".join(sorted(vals))))

    # --- il summary risponde al title? ---------------------------------------------------
    # Non si applica agli `_index`: li' il titolo e' il nome della cartella e il riassunto
    # dice cosa la cartella contiene, quindi la sovrapposizione e' bassa per costruzione e
    # l'avviso scatterebbe su tutti e undici senza dire nulla.
    if nota.type == "index":
        return
    ti = set(re.findall(r"\w{4,}", Q.norm(str(nota.fm.get("title") or ""))))
    su = set(re.findall(r"\w{4,}", Q.norm(str(nota.fm.get("summary") or ""))))
    if ti and len(ti & su) / float(len(ti)) < 0.20:
        rep.avviso(n, CONTROLLO, "summary e title si sovrappongono per meno del 20%: da ispezionare")


def controlla_locator_punta(nota, rep):
    """Riga, cella, pagina o timestamp dichiarati esistono davvero nel file."""
    for riga in nota.blocco_fonti.splitlines():
        m = re.match(r"\s*-\s*\[\[([^\]|]+)\]\]\s*(?:—|-{1,2})\s*(.+)", riga)
        if not m:
            continue
        f, loc = m.group(1).strip(), m.group(2)
        p = os.path.join(Q.SOURCES, f)
        if not os.path.isfile(p):
            continue
        est = f.rsplit(".", 1)[-1].lower()
        try:
            if est in ("log", "txt"):
                t = EC.testo_cantiere(f)
                for ts in RE_ORA.findall(loc)[:4]:
                    if ts not in t:
                        rep.errore(nota.nome, CONTROLLO,
                                   "il locator cita %s ma quel timestamp non compare in %s" % (ts, f))
            elif est == "csv":
                mm = re.search(r"riga\s+(\d+)", loc)
                if mm:
                    n_righe = EC.testo_cantiere(f).count("\n") + 1
                    if int(mm.group(1)) > n_righe:
                        rep.errore(nota.nome, CONTROLLO,
                                   "il locator cita la riga %s ma %s ne ha %d"
                                   % (mm.group(1), f, n_righe))
            elif est == "xlsx":
                from openpyxl import load_workbook
                fogli = load_workbook(p, read_only=True).sheetnames
                for nomef in re.findall(r"foglio\s+«([^»]+)»", loc):
                    if not any(Q.norm(nomef) == Q.norm(x) for x in fogli):
                        rep.errore(nota.nome, CONTROLLO,
                                   "il locator cita il foglio «%s», che in %s non esiste (ci sono: %s)"
                                   % (nomef, f, ", ".join(fogli)))
            elif est == "pdf":
                mm = re.search(r"pag\.\s*(\d+)", loc)
                if mm:
                    from pypdf import PdfReader
                    npag = len(PdfReader(p).pages)
                    if int(mm.group(1)) > npag:
                        rep.errore(nota.nome, CONTROLLO,
                                   "il locator cita pag. %s ma %s ha %d pagine"
                                   % (mm.group(1), f, npag))
            elif est == "eml":
                mm = re.search(r"header\s+([\w-]+)", loc)
                if mm and ("%s:" % mm.group(1)).lower() not in EC.testo_cantiere(f).lower():
                    rep.errore(nota.nome, CONTROLLO,
                               "il locator cita l'header %s, assente da %s" % (mm.group(1), f))
        except Exception as ex:
            rep.avviso(nota.nome, CONTROLLO,
                       "verifica del locator non eseguibile su %s: %s" % (f, str(ex)[:60]))


def pacchetto_giudizio(note, dove):
    """Prepara cio' che il subagente pulito deve ricevere: SOLO le note del lotto
    e il testo dei grezzi che citano. Niente canone, niente documenti di metodo."""
    r = [PROMPT_GIUDIZIO, "\n" + "=" * 70,
         "PROMPT %s del %s — non modificarlo per questo lotto.\\n"
         % (PROMPT_GIUDIZIO_VERSIONE, PROMPT_GIUDIZIO_DATA),
         "=" * 70 + "\n"]
    usate = []
    for n in note:
        if n.type in ("index",) or not n.fonti:
            continue
        # E10: delimitatore che non puo' comparire dentro un grezzo. Con «NOTA:» il
        # conteggio delle note inviate si falsava, perche' quella stringa compare anche
        # nel testo del manuale HACCP incluso come fonte.
        r.append("\n" + "-" * 70)
        r.append(">>>>> NOTA DA GIUDICARE: %s" % n.nome)
        r.append("-" * 70)
        r.append(n.grezzo.strip())
        usate += [str(f) for f in n.fonti]
    r.append("\n\n" + "=" * 70)
    r.append("TESTO ESTRATTO DELLE FONTI CITATE")
    r.append("=" * 70)
    # E10, SECONDO DELIMITATORE (lotto 3F, 24/08/2026). La forma «--- <nome> ---»
    # COMPARE DENTRO I GREZZI — la notifica ATS ne porta due — e chi ritaglia il
    # pacchetto la leggeva come inizio di un'altra fonte, troncando quella vera a 638
    # caratteri su 13.186. Il delimitatore delle fonti prende quindi la stessa forma di
    # quello delle note: un prefisso che in un grezzo non si scrive.
    for f in sorted(set(usate)):
        t = EC.testo_cantiere(f)
        r.append("\n>>>>> FONTE: %s" % f)
        r.append(t if t.strip() else "(nessun testo estraibile: e' un'immagine, riscontro visivo)")
    p = os.path.join(dove, "pacchetto_giudizio_provenance.txt")
    with io.open(p, "w", encoding="utf-8") as fh:
        fh.write("\n".join(r))
    return p


def main():
    ap = argparse.ArgumentParser(description="Verifica che ogni fatto abbia riscontro nelle fonti.")
    Q.aggiungi_argomenti(ap)
    ap.add_argument("--pacchetto-giudizio", action="store_true",
                    help="scrive il pacchetto per il subagente dello strato di giudizio")
    args = ap.parse_args()
    modo, file_lotto = Q.leggi_perimetro(args)

    note = Q.tutte_le_note(args.vault)
    per_slug = {n.slug: n for n in note}
    perimetro = Q.note_del_perimetro(note, modo, file_lotto, Q.note_toccate(args))

    rep = Q.Report("qa_provenance (perimetro: %s, %d note)" % (modo, len(perimetro)))
    for n in perimetro:
        controlla(n, rep, per_slug)

    rep.stampa()
    d = Q.cartella_report(args, modo, "l26130")
    Q.scrivi_report(d, "qa_provenance.md", rep.markdown())
    if args.pacchetto_giudizio:
        print("pacchetto per lo strato di giudizio: %s" % pacchetto_giudizio(perimetro, d))
    sys.exit(rep.codice())


if __name__ == "__main__":
    main()
